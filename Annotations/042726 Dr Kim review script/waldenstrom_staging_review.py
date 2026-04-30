#!/usr/bin/env python3
"""
Waldenstrom Perthes Staging Review Tool — Multi-Observer Disagreement Focus
---------------------------------------------------------------------------
Chord keys:
  1,1 → Ia    1,2 → Ib
  2,1 → IIa   2,2 → IIb
  3,1 → IIIa  3,2 → IIIb
  4   → IV    (single key)
Space → copy ipsg_waldenstrom reference to your stage

Nav:   ← →         sequential (all images)
       Page Up/Down jump between disagreement cases (filtered)

Undo:  Cmd+Z
Jump:  Cmd+G  (go to patient/filename)
Save:  continuous write-through to new Excel sheet
"""

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import os
import re
import shutil
import threading
import time
import json
import datetime
import glob
from pathlib import Path
from collections import Counter

try:
    from PIL import Image, ImageTk
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "Pillow"])
    from PIL import Image, ImageTk

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl import load_workbook

try:
    import pandas as pd
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas"])
    import pandas as pd

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

# Auto-detect paths: Excel files next to this script, OneDrive auto-discovered
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

EXCEL_SOURCE = os.path.join(SCRIPT_DIR, "042526 Waldenstrom Staging.xlsx")
EXCEL_REVIEW = os.path.join(SCRIPT_DIR, "042026 waldenstrom_staging_review_output.xlsx")
def _find_onedrive_root():
    """Auto-detect the ap_frog_pairs folder. If not found, ask the user to pick it."""
    from tkinter import filedialog
    home = Path.home()
    # Try OneDrive CloudStorage (synced)
    cloud = home / "Library" / "CloudStorage"
    if cloud.is_dir():
        for d in cloud.iterdir():
            if "scottishrite" in d.name.lower() or "onedrive" in d.name.lower():
                candidate = d / "ap_frog_pairs"
                if candidate.is_dir():
                    return str(candidate)
    # Try common locations
    for candidate in [
        home / "OneDrive - ScottishRiteforChildren" / "ap_frog_pairs",
        home / "OneDrive" / "ap_frog_pairs",
        Path(SCRIPT_DIR) / "ap_frog_pairs",
    ]:
        if candidate.is_dir():
            return str(candidate)
    # Not found — ask user to select the folder
    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo("Image Folder Needed",
                        "Could not auto-detect the ap_frog_pairs image folder.\n\n"
                        "Please select the folder containing the AP/Frog images.")
    folder = filedialog.askdirectory(title="Select ap_frog_pairs image folder")
    root.destroy()
    if folder:
        return folder
    return os.path.join(SCRIPT_DIR, "ap_frog_pairs")

ONEDRIVE_ROOT = _find_onedrive_root()
BACKUP_DIR = os.path.join(SCRIPT_DIR, "waldenstrom_backups")
AUDIT_LOG = os.path.join(SCRIPT_DIR, "waldenstrom_audit.log")
SESSION_STATE = os.path.join(SCRIPT_DIR, ".waldenstrom_session.json")

OBSERVERS = ["Joonoh", "Ahan", "Kevin", "Dr. Kim"]

CHORD_MAP = {
    "1": {"1": "Ia", "2": "Ib"},
    "2": {"1": "IIa", "2": "IIb"},
    "3": {"1": "IIIa", "2": "IIIb"},
    "4": {"4": "IV"},
}
STAGE_LEGEND = [
    ("1,1", "Ia"), ("1,2", "Ib"),
    ("2,1", "IIa"), ("2,2", "IIb"),
    ("3,1", "IIIa"), ("3,2", "IIIb"),
    ("4", "IV"),
]
CHORD_FIRST_KEYS = set(CHORD_MAP.keys())

STAGE_COLORS = {
    "Ia": "#4FC3F7", "Ib": "#0288D1",
    "IIa": "#81C784", "IIb": "#388E3C",
    "IIIa": "#FFB74D", "IIIb": "#E65100",
    "IV": "#EF5350",
    None: "#555555", "": "#555555",
}

OBSERVER_COLORS = {
    "Joonoh": "#64B5F6",
    "Ahan": "#81C784",
    "Kevin": "#FFB74D",
    "Dr. Kim": "#CE93D8",
}

TIMEPOINT_ORDER = [
    "initial", "3_month", "6_month", "12_month", "18_month", "24_month",
    "30_month", "36_month", "42_month", "48_month", "54_month",
    "5_years", "6_years", "7_years", "8_years", "9_years", "10_years",
    "final", "finalsm",
]

FILTER_OPTIONS = [
    "All Cases",
    "All 3 Agree",
    "Any Disagreement",
    "All 3 Disagree",
    "1 Observer Disagrees",
    "1 Disagrees (Joonoh)",
    "1 Disagrees (Ahan)",
    "1 Disagrees (Kevin)",
]

# ─────────────────────────────────────────────────────────────────────────────
# FILE MATCHING UTILITIES
# ─────────────────────────────────────────────────────────────────────────────

def excel_name_to_onedrive(excel_name: str) -> str | None:
    if not excel_name or not isinstance(excel_name, str):
        return None
    candidate = re.sub(r'(?i)_ap_', '_apfrog_', excel_name)
    full_path = os.path.join(ONEDRIVE_ROOT, candidate)
    if os.path.isfile(full_path):
        return full_path
    try:
        files = os.listdir(ONEDRIVE_ROOT)
    except FileNotFoundError:
        return None
    candidate_lower = candidate.lower()
    for f in files:
        if f.lower() == candidate_lower:
            return os.path.join(ONEDRIVE_ROOT, f)
    orig_lower = excel_name.lower()
    for f in files:
        if f.lower() == orig_lower:
            return os.path.join(ONEDRIVE_ROOT, f)
    return None


def extract_patient_token(filename: str) -> str:
    if not filename:
        return "UNKNOWN"
    base = os.path.splitext(os.path.basename(filename))[0]
    m = re.match(r'^(.*?)_(?:ap|AP)(?:frog)?_', base)
    if m:
        return m.group(1)
    return base.split("_")[0]


def extract_timepoint(filename: str) -> str:
    if not filename:
        return ""
    base = os.path.splitext(os.path.basename(filename))[0].lower()
    base = re.sub(r'^.*?_ap(?:frog)?_', '', base)
    return base


def timepoint_sort_key(filename: str) -> int:
    tp = extract_timepoint(filename)
    tp = re.sub(r'_[lr]$', '', tp)
    tp_lower = tp.lower()
    for i, t in enumerate(TIMEPOINT_ORDER):
        if t.lower() == tp_lower:
            return i
    return 999


def extract_laterality(filename: str) -> str:
    base = os.path.splitext(os.path.basename(filename))[0]
    if base.upper().endswith("_L"):
        return "L"
    if base.upper().endswith("_R"):
        return "R"
    return "?"

# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────────────────────────────────────

def load_consolidated_excel(path: str) -> pd.DataFrame:
    try:
        return pd.read_excel(path, sheet_name="Consolidated")
    except ValueError:
        # If "Consolidated" sheet not found, try the first sheet
        return pd.read_excel(path)


def ensure_review_file():
    """Create the review output file from the source if it doesn't exist yet."""
    if not os.path.isfile(EXCEL_REVIEW):
        wb = openpyxl.Workbook()
        # Do not remove the default sheet here; it will be removed when saving actual data
        wb.save(EXCEL_REVIEW)
    return EXCEL_REVIEW


def load_existing_review_sheet(observer: str) -> tuple[dict, str | None]:
    """Check review output file for existing Review sheets for this observer."""
    staging = {}
    sheet_name = None
    if not os.path.isfile(EXCEL_REVIEW):
        return staging, sheet_name
    try:
        wb = load_workbook(EXCEL_REVIEW, read_only=True, data_only=True)
        review_sheets = [s for s in wb.sheetnames if s.startswith("Review_")]
        review_sheets.sort(reverse=True)
        for sn in review_sheets:
            ws = wb[sn]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            obs_col_name = f"{observer}_review"
            if obs_col_name in headers:
                obs_idx = headers.index(obs_col_name)
                fn_idx = headers.index("ap_file_name")
                for row in ws.iter_rows(min_row=2, values_only=True):
                    fn = row[fn_idx]
                    val = row[obs_idx]
                    if fn and val:
                        staging[str(fn).strip()] = str(val).strip()
                sheet_name = sn
                break
        wb.close()
    except Exception:
        pass
    return staging, sheet_name


def build_image_list(df: pd.DataFrame, ipsg_ref: dict) -> list[dict]:
    valid = []
    for _, row in df.iterrows():
        fn = str(row.get("ap_file_name", "")).strip()
        if not fn or fn == "nan":
            continue
        valid.append({
            "ap_file_name": fn,
            "patient_token": str(row.get("patient_token", "")),
            "laterality": str(row.get("laterality", "")),
            "Joonoh": str(row.get("Joonoh", "")) if pd.notna(row.get("Joonoh")) else None,
            "Ahan": str(row.get("Ahan", "")) if pd.notna(row.get("Ahan")) else None,
            "Kevin": str(row.get("Kevin", "")) if pd.notna(row.get("Kevin")) else None,
            "Dr. Kim": str(row.get("Dr. Kim", "")) if pd.notna(row.get("Dr. Kim", None)) else None,
            "disagreement": str(row.get("Disagreement", "")) if pd.notna(row.get("Disagreement")) else "",
            "ipsg_waldenstrom": ipsg_ref.get(fn, ""),
            "onedrive_path": excel_name_to_onedrive(fn),
        })

    groups: dict[tuple, list] = {}
    for item in valid:
        fn = item["ap_file_name"]
        pt = item["patient_token"]
        lat = item["laterality"]
        key = (pt, lat)
        groups.setdefault(key, []).append(item)

    for key in groups:
        groups[key].sort(key=lambda r: timepoint_sort_key(r["ap_file_name"]))

    def group_sort_key(k):
        token = k[0]
        m = re.search(r'(\d+)', token)
        num = int(m.group(1)) if m else 0
        return (token.lower().startswith("patient_"), num, token.lower(), k[1])

    sorted_keys = sorted(groups.keys(), key=group_sort_key)
    result = []
    for key in sorted_keys:
        for item in groups[key]:
            result.append(item)
    return result

# ─────────────────────────────────────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────────────────────────────────────

_save_lock = threading.Lock()


def save_to_review_sheet(image_list: list[dict],
                         staging: dict, observer: str, sheet_name: str):
    """Save current staging to a review sheet in the separate review output file."""
    ensure_review_file()
    try:
        wb = load_workbook(EXCEL_REVIEW)
    except Exception:
        wb = openpyxl.Workbook()
        # Remove default empty sheet if present
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    obs_col = f"{observer}_review"
    headers = ["ap_file_name", "patient_token", "laterality",
               "Joonoh", "Ahan", "Kevin", "Disagreement", obs_col]
    ws.append(headers)

    for img in image_list:
        fn = img["ap_file_name"]
        review_val = staging.get(fn, "") or ""
        row_data = [
            fn,
            img["patient_token"],
            img["laterality"],
            img.get("Joonoh") or "",
            img.get("Ahan") or "",
            img.get("Kevin") or "",
            img.get("disagreement") or "",
            review_val,
        ]
        ws.append(row_data)

    wb.save(EXCEL_REVIEW)


def write_audit(filename: str, old_val, new_val: str, observer: str):
    ts = datetime.datetime.now().isoformat()
    with open(AUDIT_LOG, 'a', encoding='utf-8') as f:
        f.write(f"{ts}\t{observer}\t{filename}\t{old_val!r}\t{new_val!r}\n")


def save_session_state(index: int, observer: str, filter_mode: str, sheet_name: str):
    with open(SESSION_STATE, 'w') as f:
        json.dump({"index": index, "observer": observer,
                   "filter_mode": filter_mode, "sheet_name": sheet_name,
                   "ts": time.time()}, f)


def load_session_state() -> dict | None:
    if not os.path.isfile(SESSION_STATE):
        return None
    try:
        with open(SESSION_STATE) as f:
            return json.load(f)
    except Exception:
        return None

# ─────────────────────────────────────────────────────────────────────────────
# OBSERVER SELECTION DIALOG
# ─────────────────────────────────────────────────────────────────────────────

def ask_observer(last_observer=None):
    """Show a standalone Tk dialog to pick the observer. Returns name or None."""
    root = tk.Tk()
    root.title("Select Observer")
    root.configure(bg="#1a1a2e")
    root.resizable(False, False)
    root.geometry("320x330")

    result = [None]
    var = tk.StringVar(value=last_observer or OBSERVERS[0])

    tk.Label(root, text="Select which observer\nyou are reviewing for:",
             bg="#1a1a2e", fg="#7ec8e3", font=("Menlo", 14, "bold"),
             justify="center").pack(padx=30, pady=(20, 15))

    for obs in OBSERVERS:
        color = OBSERVER_COLORS.get(obs, "#ffffff")
        rb = tk.Radiobutton(root, text=obs, variable=var, value=obs,
                            bg="#16213e", fg=color, selectcolor="#0f3460",
                            activebackground="#16213e", activeforeground=color,
                            font=("Menlo", 13), indicatoron=True, anchor="w",
                            padx=20)
        rb.pack(fill="x", padx=30, pady=3)

    def _ok():
        result[0] = var.get()
        root.destroy()

    def _cancel():
        result[0] = None
        root.destroy()

    btn_frame = tk.Frame(root, bg="#1a1a2e")
    btn_frame.pack(pady=(15, 20))
    tk.Button(btn_frame, text="Start Session", command=_ok,
              bg="#0f3460", fg="white", font=("Menlo", 12, "bold"),
              relief="flat", padx=20, pady=6, cursor="hand2",
              activebackground="#1a5276").pack()

    root.protocol("WM_DELETE_WINDOW", _cancel)
    root.mainloop()
    return result[0]

# ─────────────────────────────────────────────────────────────────────────────
# MAIN APPLICATION
# ─────────────────────────────────────────────────────────────────────────────

class StagingApp(tk.Tk):
    def __init__(self, active_observer: str):
        super().__init__()
        self.active_observer = active_observer
        self.title(f"Waldenstrom Staging Review — {self.active_observer}")
        self.configure(bg="#1a1a2e")
        # Use geometry instead of state("zoomed") for macOS compatibility
        self.geometry("1400x900")
        try:
            self.state("zoomed")
        except Exception:
            pass

        # Load data from source (read-only)
        print("Loading consolidated Excel (read-only source)...")
        df = load_consolidated_excel(EXCEL_SOURCE)
        print(f"Loaded {len(df)} rows")
        print("Building image list...")
        self.image_list = build_image_list(df, {})
        print(f"Built {len(self.image_list)} images")

        # Ensure review output file exists
        ensure_review_file()

        # Load existing review or start fresh
        session = load_session_state()
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        self.review_sheet_name = f"Review_{ts}"

        # Check if resuming
        self.staging: dict[str, str | None] = {}
        if session and session.get("observer") == self.active_observer and session.get("sheet_name"):
            existing, found_sheet = load_existing_review_sheet(self.active_observer)
            if existing and found_sheet:
                if messagebox.askyesno("Resume?",
                                       f"Found previous review session ({found_sheet}) "
                                       f"for {self.active_observer} with {len(existing)} entries.\n\n"
                                       f"Resume from that session?"):
                    self.staging = existing
                    self.review_sheet_name = found_sheet

        self._undo_stack: list[tuple] = []
        self._img_pil_map: dict[str, Image.Image] = {}

        # Filter state
        default_filter = "Any Disagreement"
        if session and session.get("filter_mode"):
            default_filter = session["filter_mode"]
        self.filter_mode = tk.StringVar(value=default_filter)
        self.filtered_indices: list[int] = []
        self._rebuild_filtered()

        # Start index
        saved_idx = session.get("index") if session else None
        if saved_idx is not None and 0 <= saved_idx < len(self.image_list):
            self.current_index = saved_idx
        else:
            self.current_index = self.filtered_indices[0] if self.filtered_indices else 0

        # Chord input state
        self._chord_first: str | None = None
        self._chord_timer: str | None = None

        # Build UI
        self._build_ui()

        # Key bindings
        self.bind("<Left>", lambda e: self._navigate(-1))
        self.bind("<Right>", lambda e: self._navigate(1))
        self.bind("<Prior>", lambda e: self._jump_disagreement(-1))  # Page Up
        self.bind("<Next>", lambda e: self._jump_disagreement(1))    # Page Down
        self.bind("<Command-z>", lambda e: self._undo())
        self.bind("<Command-g>", lambda e: self._jump_dialog())
        self.bind("<space>", lambda e: self._copy_reference())
        for k in CHORD_FIRST_KEYS:
            self.bind(k, lambda e, key=k: self._handle_chord_key(key))

        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self._load_current()

    # ── Filter ────────────────────────────────────────────────────────────────

    def _rebuild_filtered(self):
        mode = self.filter_mode.get()
        self.filtered_indices = []
        for i, img in enumerate(self.image_list):
            dis = img.get("disagreement", "")
            if mode == "All Cases":
                self.filtered_indices.append(i)
            elif mode == "All 3 Agree":
                if "All 3 Agree" in dis or "Agree" in dis:
                    self.filtered_indices.append(i)
            elif mode == "Any Disagreement":
                if "Disagree" in dis:
                    self.filtered_indices.append(i)
            elif mode == "All 3 Disagree":
                if "All 3 Disagree" in dis:
                    self.filtered_indices.append(i)
            elif mode == "1 Observer Disagrees":
                if "1 Disagrees" in dis:
                    self.filtered_indices.append(i)
            elif mode == "1 Disagrees (Joonoh)":
                if "1 Disagrees (Joonoh)" in dis:
                    self.filtered_indices.append(i)
            elif mode == "1 Disagrees (Ahan)":
                if "1 Disagrees (Ahan)" in dis:
                    self.filtered_indices.append(i)
            elif mode == "1 Disagrees (Kevin)":
                if "1 Disagrees (Kevin)" in dis:
                    self.filtered_indices.append(i)

    def _on_filter_change(self, *args):
        self._rebuild_filtered()
        count = len(self.filtered_indices)
        self._filter_count_label.config(text=f"{count} cases match filter")

    # ── UI Construction ──────────────────────────────────────────────────────

    def _build_ui(self):
        self.columnconfigure(0, weight=0)
        self.columnconfigure(1, weight=1)
        self.columnconfigure(2, weight=0)
        self.rowconfigure(0, weight=1)

        # ── Left sidebar ──
        left = tk.Frame(self, bg="#16213e", width=260)
        left.grid(row=0, column=0, sticky="nsew", padx=(8, 0), pady=8)
        left.grid_propagate(False)

        # Observer banner
        obs_color = OBSERVER_COLORS.get(self.active_observer, "#ffffff")
        tk.Label(left, text=f"REVIEWING AS: {self.active_observer.upper()}",
                 bg=obs_color, fg="#000000",
                 font=("Menlo", 10, "bold"), anchor="center"
                 ).pack(fill="x", padx=6, pady=(6, 4))

        # Disagreement filter
        tk.Label(left, text="DISAGREEMENT FILTER", bg="#16213e", fg="#7ec8e3",
                 font=("Menlo", 9, "bold"), anchor="w").pack(fill="x", padx=10, pady=(8, 2))

        filter_frame = tk.Frame(left, bg="#16213e")
        filter_frame.pack(fill="x", padx=6)
        for opt in FILTER_OPTIONS:
            rb = tk.Radiobutton(filter_frame, text=opt, variable=self.filter_mode,
                                value=opt, bg="#16213e", fg="#cccccc",
                                selectcolor="#0f3460", activebackground="#16213e",
                                activeforeground="#ffffff", font=("Menlo", 9),
                                anchor="w", indicatoron=True,
                                command=self._on_filter_change)
            rb.pack(fill="x", padx=4, pady=0)

        self._filter_count_label = tk.Label(left, text="", bg="#16213e", fg="#888888",
                                            font=("Menlo", 8))
        self._filter_count_label.pack(fill="x", padx=10, pady=(2, 6))
        self._on_filter_change()

        sep = tk.Frame(left, bg="#2a3a5e", height=1)
        sep.pack(fill="x", padx=6, pady=4)

        # Patient timeline
        tk.Label(left, text="PATIENT SERIES", bg="#16213e", fg="#7ec8e3",
                 font=("Menlo", 9, "bold"), anchor="w").pack(fill="x", padx=10, pady=(4, 2))

        timeline_container = tk.Frame(left, bg="#16213e")
        timeline_container.pack(fill="both", expand=True, padx=6, pady=2)

        self._timeline_canvas = tk.Canvas(timeline_container, bg="#16213e",
                                          highlightthickness=0)
        self._timeline_scroll = tk.Scrollbar(timeline_container, orient="vertical",
                                             command=self._timeline_canvas.yview)
        self._timeline_frame = tk.Frame(self._timeline_canvas, bg="#16213e")
        self._timeline_frame.bind("<Configure>",
            lambda e: self._timeline_canvas.configure(scrollregion=self._timeline_canvas.bbox("all")))
        self._timeline_canvas.create_window((0, 0), window=self._timeline_frame, anchor="nw")
        self._timeline_canvas.configure(yscrollcommand=self._timeline_scroll.set)
        self._timeline_canvas.pack(side="left", fill="both", expand=True)
        self._timeline_scroll.pack(side="right", fill="y")

        # Navigation
        nav = tk.Frame(left, bg="#16213e")
        nav.pack(fill="x", padx=6, pady=4)
        tk.Button(nav, text="◀ Prev", command=lambda: self._navigate(-1),
                  bg="#0f3460", fg="white", relief="flat", font=("Menlo", 10),
                  activebackground="#1a5276", cursor="hand2"
                  ).pack(side="left", expand=True, fill="x", padx=(0, 2))
        tk.Button(nav, text="Next ▶", command=lambda: self._navigate(1),
                  bg="#0f3460", fg="white", relief="flat", font=("Menlo", 10),
                  activebackground="#1a5276", cursor="hand2"
                  ).pack(side="left", expand=True, fill="x", padx=(2, 0))

        # Disagreement jump buttons
        djnav = tk.Frame(left, bg="#16213e")
        djnav.pack(fill="x", padx=6, pady=(2, 4))
        tk.Button(djnav, text="⏮ Prev Disagree", command=lambda: self._jump_disagreement(-1),
                  bg="#5c2a0f", fg="#FFB74D", relief="flat", font=("Menlo", 9),
                  activebackground="#7a3a1a", cursor="hand2"
                  ).pack(side="left", expand=True, fill="x", padx=(0, 2))
        tk.Button(djnav, text="Next Disagree ⏭", command=lambda: self._jump_disagreement(1),
                  bg="#5c2a0f", fg="#FFB74D", relief="flat", font=("Menlo", 9),
                  activebackground="#7a3a1a", cursor="hand2"
                  ).pack(side="left", expand=True, fill="x", padx=(2, 0))

        tk.Button(left, text="Cmd+G  Jump to Patient", command=self._jump_dialog,
                  bg="#0f3460", fg="#7ec8e3", relief="flat", font=("Menlo", 9),
                  activebackground="#1a5276", cursor="hand2"
                  ).pack(fill="x", padx=6, pady=(0, 4))

        self._progress_label = tk.Label(left, text="", bg="#16213e", fg="#aaaaaa",
                                        font=("Menlo", 8), anchor="w")
        self._progress_label.pack(fill="x", padx=10, pady=(0, 6))

        # ── Center: image ──
        center = tk.Frame(self, bg="#1a1a2e")
        center.grid(row=0, column=1, sticky="nsew", padx=8, pady=8)
        center.rowconfigure(0, weight=1)
        center.rowconfigure(1, weight=0)
        center.columnconfigure(0, weight=1)

        self._canvas = tk.Canvas(center, bg="#0d0d0d", highlightthickness=0)
        self._canvas.grid(row=0, column=0, sticky="nsew")
        self._canvas.bind("<Configure>", lambda e: self._redraw_image())

        self._fname_label = tk.Label(center, text="", bg="#1a1a2e", fg="#cccccc",
                                     font=("Menlo", 9), wraplength=800)
        self._fname_label.grid(row=1, column=0, pady=(4, 4))

        # ── Right panel ──
        right = tk.Frame(self, bg="#16213e", width=280)
        right.grid(row=0, column=2, sticky="nsew", padx=(0, 8), pady=8)
        right.grid_propagate(False)
        right.columnconfigure(0, weight=1)

        # YOUR REVIEW STAGE (large)
        tk.Label(right, text="YOUR REVIEW", bg="#16213e", fg="#7ec8e3",
                 font=("Menlo", 10, "bold"), anchor="w").pack(fill="x", padx=10, pady=(10, 2))

        self._stage_display = tk.Label(right, text="—", bg="#16213e", fg="#ffffff",
                                       font=("Menlo", 44, "bold"))
        self._stage_display.pack(pady=(0, 6))

        # ALL OBSERVERS section
        sep1 = tk.Frame(right, bg="#2a3a5e", height=1)
        sep1.pack(fill="x", padx=10, pady=4)

        tk.Label(right, text="ALL OBSERVERS", bg="#16213e", fg="#7ec8e3",
                 font=("Menlo", 10, "bold"), anchor="w").pack(fill="x", padx=10, pady=(4, 4))

        self._observer_frames = {}
        self._observer_labels = {}
        for obs in OBSERVERS:
            obs_frame = tk.Frame(right, bg="#1a2a4a", relief="flat")
            obs_frame.pack(fill="x", padx=10, pady=2)

            obs_color = OBSERVER_COLORS.get(obs, "#ffffff")
            # Active observer indicator
            indicator = "▶ " if obs == self.active_observer else "  "
            name_label = tk.Label(obs_frame, text=f"{indicator}{obs}",
                                  bg="#1a2a4a", fg=obs_color,
                                  font=("Menlo", 11, "bold"), anchor="w", width=12)
            name_label.pack(side="left", padx=(6, 4), pady=4)

            stage_label = tk.Label(obs_frame, text="—", bg="#1a2a4a", fg="#ffffff",
                                   font=("Menlo", 18, "bold"), anchor="center", width=5)
            stage_label.pack(side="right", padx=(4, 8), pady=4)

            self._observer_frames[obs] = obs_frame
            self._observer_labels[obs] = stage_label

        # Disagreement status
        sep2 = tk.Frame(right, bg="#2a3a5e", height=1)
        sep2.pack(fill="x", padx=10, pady=6)

        self._disagree_label = tk.Label(right, text="", bg="#16213e",
                                        font=("Menlo", 10, "bold"), wraplength=240)
        self._disagree_label.pack(fill="x", padx=10, pady=(0, 4))

        # REFERENCE (ipsg)
        sep3 = tk.Frame(right, bg="#2a3a5e", height=1)
        sep3.pack(fill="x", padx=10, pady=4)

        tk.Label(right, text="REFERENCE (ipsg)", bg="#16213e", fg="#7ec8e3",
                 font=("Menlo", 10, "bold"), anchor="w").pack(fill="x", padx=10, pady=(4, 2))

        self._ref_display = tk.Label(right, text="—", bg="#16213e", fg="#aaaaaa",
                                     font=("Menlo", 28, "bold"))
        self._ref_display.pack(pady=(0, 6))

        # Keys legend
        sep4 = tk.Frame(right, bg="#2a3a5e", height=1)
        sep4.pack(fill="x", padx=10, pady=4)

        tk.Label(right, text="KEYS", bg="#16213e", fg="#7ec8e3",
                 font=("Menlo", 9, "bold"), anchor="w").pack(fill="x", padx=10, pady=(4, 2))

        legend_frame = tk.Frame(right, bg="#16213e")
        legend_frame.pack(fill="x", padx=10)
        for chord, stage in STAGE_LEGEND:
            color = STAGE_COLORS.get(stage, "#888888")
            row_f = tk.Frame(legend_frame, bg="#16213e")
            row_f.pack(fill="x", pady=0)
            tk.Label(row_f, text=f" {chord}", bg="#16213e", fg="#888888",
                     font=("Menlo", 10), width=4, anchor="w").pack(side="left")
            tk.Label(row_f, text="→", bg="#16213e", fg="#444", font=("Menlo", 10)
                     ).pack(side="left")
            tk.Label(row_f, text=f" {stage}", bg="#16213e", fg=color,
                     font=("Menlo", 10, "bold")).pack(side="left")

        # Chord pending
        self._chord_label = tk.Label(right, text="", bg="#16213e",
                                     fg="#FFD700", font=("Menlo", 12, "bold"))
        self._chord_label.pack(pady=(4, 2))

        tk.Label(right, text="Cmd+Z undo  Cmd+G jump\n←/→ sequential  PgUp/Dn disagree\nSPACE copy reference",
                 bg="#16213e", fg="#555555", font=("Menlo", 8)).pack(pady=(0, 6))

        self._counter_label = tk.Label(right, text="", bg="#16213e", fg="#555",
                                       font=("Menlo", 8))
        self._counter_label.pack(pady=(0, 6))

    # ── Image Loading ────────────────────────────────────────────────────────

    def _load_current(self):
        if not self.image_list:
            return
        img_data = self.image_list[self.current_index]
        path = img_data.get("onedrive_path")
        fn = img_data["ap_file_name"]

        self._fname_label.config(text=fn)

        if path and os.path.isfile(path):
            self._load_image(path)
        else:
            self._canvas.delete("all")
            w = self._canvas.winfo_width() or 800
            h = self._canvas.winfo_height() or 600
            self._canvas.create_text(w // 2, h // 2, text="Image not found\n" + fn,
                                     fill="#ff4444", font=("Menlo", 14), justify="center")
            self._current_path = None

        # Update YOUR REVIEW stage
        review_stage = self.staging.get(fn)
        self._stage_display.config(
            text=review_stage or "—",
            fg=STAGE_COLORS.get(review_stage, "#555555")
        )

        # Update ALL OBSERVERS
        for obs in OBSERVERS:
            orig_stage = img_data.get(obs)
            label = self._observer_labels[obs]
            frame = self._observer_frames[obs]

            display_val = orig_stage or "—"
            color = STAGE_COLORS.get(orig_stage, "#555555")
            label.config(text=display_val, fg=color)

            # Highlight disagreement: compare to majority (original 3 observers only)
            orig_3 = ["Joonoh", "Ahan", "Kevin"]
            stages = [img_data.get(o) for o in orig_3 if img_data.get(o)]
            if len(stages) >= 2:
                counts = Counter(stages)
                majority = counts.most_common(1)[0][0]
                if orig_stage and orig_stage != majority:
                    frame.config(bg="#3a1a1a")  # red tint for dissenter
                    label.config(bg="#3a1a1a")
                    # Also update the name label
                    for child in frame.winfo_children():
                        child.config(bg="#3a1a1a")
                else:
                    frame.config(bg="#1a2a4a")
                    label.config(bg="#1a2a4a")
                    for child in frame.winfo_children():
                        child.config(bg="#1a2a4a")
            else:
                frame.config(bg="#1a2a4a")
                label.config(bg="#1a2a4a")
                for child in frame.winfo_children():
                    child.config(bg="#1a2a4a")

        # Disagreement label
        dis = img_data.get("disagreement", "")
        if "All 3 Disagree" in dis:
            self._disagree_label.config(text=f"ALL 3 DISAGREE", fg="#EF5350")
        elif "1 Disagrees" in dis:
            self._disagree_label.config(text=dis, fg="#FFB74D")
        elif "All 3 Agree" in dis:
            self._disagree_label.config(text="All 3 Agree", fg="#4CAF50")
        else:
            self._disagree_label.config(text=dis or "—", fg="#888888")

        # Reference
        ref = img_data.get("ipsg_waldenstrom", "") or ""
        self._ref_display.config(text=ref or "—")

        # Counter
        total = len(self.image_list)
        staged = sum(1 for v in self.staging.values() if v)
        dis_idx = -1
        if self.current_index in self.filtered_indices:
            dis_idx = self.filtered_indices.index(self.current_index) + 1
        filter_total = len(self.filtered_indices)
        filter_text = f"  |  Filtered: {dis_idx}/{filter_total}" if dis_idx > 0 else f"  |  Filter: {filter_total} cases"
        self._counter_label.config(
            text=f"Image {self.current_index + 1}/{total}  ({staged} reviewed){filter_text}")

        self._progress_label.config(
            text=f"{staged}/{total} reviewed ({100 * staged // max(total, 1)}%)")

        self._update_timeline()
        save_session_state(self.current_index, self.active_observer,
                           self.filter_mode.get(), self.review_sheet_name)
        self.after(50, self._preload_adjacent)

    def _load_image(self, path: str):
        self._current_path = path
        if path not in self._img_pil_map:
            try:
                self._img_pil_map[path] = Image.open(path)
            except Exception as e:
                self._canvas.delete("all")
                w = self._canvas.winfo_width() or 800
                h = self._canvas.winfo_height() or 600
                self._canvas.create_text(w // 2, h // 2, text=f"Error: {e}",
                                         fill="#ff4444", font=("Menlo", 12))
                return
        self._redraw_image()

    def _redraw_image(self):
        path = getattr(self, '_current_path', None)
        if not path or path not in self._img_pil_map:
            return
        pil_img = self._img_pil_map[path]
        cw = self._canvas.winfo_width()
        ch = self._canvas.winfo_height()
        if cw < 10 or ch < 10:
            return
        iw, ih = pil_img.size
        scale = min(cw / iw, ch / ih)
        nw, nh = int(iw * scale), int(ih * scale)
        resized = pil_img.resize((nw, nh), Image.LANCZOS)
        photo = ImageTk.PhotoImage(resized)
        self._canvas.delete("all")
        self._canvas.create_image(cw // 2, ch // 2, anchor="center", image=photo)
        self._canvas.image = photo

    def _preload_adjacent(self):
        indices = [self.current_index + d for d in (-2, -1, 1, 2)
                   if 0 <= self.current_index + d < len(self.image_list)]
        for i in indices:
            path = self.image_list[i].get("onedrive_path")
            if path and os.path.isfile(path) and path not in self._img_pil_map:
                try:
                    self._img_pil_map[path] = Image.open(path)
                except Exception:
                    pass

    # ── Timeline ─────────────────────────────────────────────────────────────

    def _update_timeline(self):
        for w in self._timeline_frame.winfo_children():
            w.destroy()

        current = self.image_list[self.current_index]
        pt = current["patient_token"]
        lat = current["laterality"]

        series = [img for img in self.image_list
                  if img["patient_token"] == pt and img["laterality"] == lat]

        tk.Label(self._timeline_frame,
                 text=f"{pt} · {lat}", bg="#16213e", fg="#7ec8e3",
                 font=("Menlo", 9, "bold"), anchor="w").pack(fill="x", pady=(0, 4))

        for img in series:
            fn = img["ap_file_name"]
            idx = self.image_list.index(img)
            is_current = (idx == self.current_index)

            tp = extract_timepoint(fn)
            tp = re.sub(r'_[lr]$', '', tp, flags=re.IGNORECASE)

            bg = "#0f3460" if is_current else "#1a2a4a"

            row_f = tk.Frame(self._timeline_frame, bg=bg, cursor="hand2")
            row_f.pack(fill="x", pady=1)
            row_f.bind("<Button-1>", lambda e, i=idx: self._jump_to(i))

            # Timepoint label
            tp_label = tk.Label(row_f, text=tp[:12], bg=bg,
                                fg="#cccccc" if is_current else "#888",
                                font=("Menlo", 8), anchor="w", width=12)
            tp_label.pack(side="left", padx=(4, 2))
            tp_label.bind("<Button-1>", lambda e, i=idx: self._jump_to(i))

            # Mini observer stages: J A K D
            for obs_short, obs_full in [("J", "Joonoh"), ("A", "Ahan"), ("K", "Kevin"), ("D", "Dr. Kim")]:
                stage = img.get(obs_full)
                color = STAGE_COLORS.get(stage, "#333")
                display = stage[:3] if stage else "·"
                lbl = tk.Label(row_f, text=display, bg=bg, fg=color,
                               font=("Menlo", 7, "bold"), width=4, anchor="center")
                lbl.pack(side="left", padx=0)
                lbl.bind("<Button-1>", lambda e, i=idx: self._jump_to(i))

            # Review stage if any
            rev = self.staging.get(fn)
            if rev:
                rev_lbl = tk.Label(row_f, text=f"→{rev}", bg=bg,
                                   fg="#FFD700", font=("Menlo", 7, "bold"),
                                   anchor="e")
                rev_lbl.pack(side="right", padx=(0, 4))
                rev_lbl.bind("<Button-1>", lambda e, i=idx: self._jump_to(i))

    # ── Staging Actions ──────────────────────────────────────────────────────

    def _handle_chord_key(self, key: str):
        if self._chord_timer is not None:
            self.after_cancel(self._chord_timer)
            self._chord_timer = None

        if self._chord_first is None:
            sub = CHORD_MAP.get(key, {})
            if key == "4":
                self._chord_first = None
                self._chord_label.config(text="")
                self._apply_stage("IV")
            elif len(sub) == 1:
                self._chord_first = None
                self._chord_label.config(text="")
                self._apply_stage(list(sub.values())[0])
            else:
                self._chord_first = key
                self._chord_label.config(text=f"  {key} + ?  (1 or 2)")
                self._chord_timer = self.after(1500, self._chord_timeout)
        else:
            first = self._chord_first
            self._chord_first = None
            self._chord_label.config(text="")
            sub = CHORD_MAP.get(first, {})
            stage = sub.get(key)
            if stage:
                self._apply_stage(stage)
            else:
                self._handle_chord_key(key)

    def _chord_timeout(self):
        self._chord_first = None
        self._chord_timer = None
        self._chord_label.config(text="")

    def _apply_stage(self, stage: str):
        fn = self.image_list[self.current_index]["ap_file_name"]
        old = self.staging.get(fn)
        self._undo_stack.append((self.current_index, fn, old))
        if len(self._undo_stack) > 200:
            self._undo_stack.pop(0)

        self.staging[fn] = stage
        write_audit(fn, old, stage, self.active_observer)

        # Save to review output file in background
        def _bg_save():
            with _save_lock:
                save_to_review_sheet(self.image_list,
                                     self.staging, self.active_observer,
                                     self.review_sheet_name)
        threading.Thread(target=_bg_save, daemon=True).start()

        self._stage_display.config(text=stage, fg=STAGE_COLORS.get(stage, "white"))
        self._flash(stage)
        self._navigate(1)

    def _copy_reference(self):
        img_data = self.image_list[self.current_index]
        ref = (img_data.get("ipsg_waldenstrom", "") or "").strip()
        if not ref:
            self._canvas.config(bg="#440000")
            self.after(150, lambda: self._canvas.config(bg="#0d0d0d"))
            return
        self._apply_stage(ref)

    def _flash(self, stage: str):
        color = STAGE_COLORS.get(stage, "#ffffff")
        self._canvas.config(bg=color)
        self.after(80, lambda: self._canvas.config(bg="#0d0d0d"))

    def _undo(self):
        if not self._undo_stack:
            return
        idx, fn, old_val = self._undo_stack.pop()
        current_val = self.staging.get(fn)
        self.staging[fn] = old_val
        write_audit(fn, current_val, old_val, self.active_observer)

        def _bg_save():
            with _save_lock:
                save_to_review_sheet(self.image_list,
                                     self.staging, self.active_observer,
                                     self.review_sheet_name)
        threading.Thread(target=_bg_save, daemon=True).start()

        self.current_index = idx
        self._load_current()

    # ── Navigation ───────────────────────────────────────────────────────────

    def _navigate(self, delta: int):
        new_idx = self.current_index + delta
        if 0 <= new_idx < len(self.image_list):
            self.current_index = new_idx
            self._load_current()

    def _jump_disagreement(self, delta: int):
        if not self.filtered_indices:
            return
        # Find next/prev in filtered list
        if delta > 0:
            for fi in self.filtered_indices:
                if fi > self.current_index:
                    self.current_index = fi
                    self._load_current()
                    return
            # Wrap to first
            self.current_index = self.filtered_indices[0]
            self._load_current()
        else:
            for fi in reversed(self.filtered_indices):
                if fi < self.current_index:
                    self.current_index = fi
                    self._load_current()
                    return
            # Wrap to last
            self.current_index = self.filtered_indices[-1]
            self._load_current()

    def _jump_to(self, idx: int):
        self.current_index = idx
        self._load_current()

    def _jump_dialog(self):
        query = simpledialog.askstring(
            "Jump to Patient",
            "Enter patient token or partial filename:",
            parent=self
        )
        if not query:
            return
        query_l = query.strip().lower()
        for i, img in enumerate(self.image_list):
            if (query_l in img["patient_token"].lower() or
                    query_l in img["ap_file_name"].lower()):
                self.current_index = i
                self._load_current()
                return
        messagebox.showinfo("Not found", f"No match for: {query}")

    # ── Close ────────────────────────────────────────────────────────────────

    def _on_close(self):
        with _save_lock:
            save_to_review_sheet(self.image_list,
                                 self.staging, self.active_observer,
                                 self.review_sheet_name)
        save_session_state(self.current_index, self.active_observer,
                           self.filter_mode.get(), self.review_sheet_name)
        self.destroy()


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    os.makedirs(BACKUP_DIR, exist_ok=True)

    # Observer selection (standalone dialog before main app)
    session = load_session_state()
    last_obs = session.get("observer") if session else None
    observer = ask_observer(last_observer=last_obs)
    if not observer:
        print("No observer selected. Exiting.")
        exit(0)

    print(f"Starting session for: {observer}")
    try:
        app = StagingApp(active_observer=observer)
        app.mainloop()
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"\nERROR: {e}")
        input("Press Enter to exit...")
