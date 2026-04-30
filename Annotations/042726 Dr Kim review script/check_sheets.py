import openpyxl

try:
    wb = openpyxl.load_workbook('042526 Waldenstrom Staging.xlsx')
    print('Sheets:', wb.sheetnames)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Sheet '{sheet_name}' has {ws.max_row} rows and {ws.max_column} columns")
except Exception as e:
    print(f"Error: {e}")
